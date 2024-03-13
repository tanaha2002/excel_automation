# %%
import openpyxl
import datetime
import excelautopd
import pandas as pd
from openpyxl.styles import Alignment, Border
from logger import logfile
# from configs import xlsx_report, xlsx_data, date_sheets

# %%

def extract_data(path, sheet_name):
    df = excelautopd.get_data(path, sheet_name)
    df_debit = df[['Project','Supplier', 'Description', 'Total', 'Payment', 'Remain']].copy()
    df_debit.loc[:, 'Title'] = df['Description'].apply(lambda x: x.split('\n')[0])
    df_debit.loc[:,'Date'] = sheet_name
    return df_debit

def get_supplier_positions(sheet):
    supplier_positions = {}
    for row_index, row in enumerate(sheet.iter_rows(min_row=1, min_col=1, max_col=1, values_only=True), start=1):
        supplier = row[0]
        supplier_positions[row_index] = supplier
    return supplier_positions

def append_data(path):  
    sheet_debit = 'DEBIT'
    workbook = openpyxl.load_workbook(filename=path)
    sheet = workbook[sheet_debit]
    # Store existing supplier data for faster lookup
    supplier_positions = get_supplier_positions(sheet)
    # Precalculate 'total' row
    for stt, row in enumerate(sheet.iter_rows(min_row=1, min_col=3, max_col=3), start=1):
        total_row = stt
    total_row +=1
    print(supplier_positions)
    return workbook, sheet, supplier_positions, total_row

def eddit_row_data(sheet, insert_row_idx, data,same_supplier=False):
    old_remain = sheet.cell(row=insert_row_idx, column=6).value
    sheet.cell(row=insert_row_idx, column=3, value=data['Description'])
    sheet.cell(row=insert_row_idx, column=4, value=data['Total'])
    sheet.cell(row=insert_row_idx, column=5, value=data['Payment'])
    sheet.cell(row=insert_row_idx, column=6, value=data['Remain'])
    sheet.cell(row=insert_row_idx, column=7, value=data['Project'])
    sheet.cell(row=insert_row_idx, column=8, value=data['Title'])
    sheet.cell(row=insert_row_idx, column=9, value=data['Date'])
    #Apply auto fit row height
    #Autofit the row height
    sheet.row_dimensions[insert_row_idx].height = None
    #get merge range
    merged_ranges = list(sheet.merged_cells.ranges)  # Get merged cells
    #get the merge range that contain insert_row_idx
    for merged_range in merged_ranges:
        if insert_row_idx >= merged_range.min_row and insert_row_idx <= merged_range.max_row:
            if merged_range.min_col == 2:
                sheet.cell(row=merged_range.min_row, column=2, value=sheet.cell(row=merged_range.min_row, column=2).value - old_remain + data['Remain']) #Column 2
                break

def insert_row_data(sheet, insert_row_idx, data,same_supplier=False):
    # Identify merged cells
    merged_ranges = list(sheet.merged_cells.ranges)  
    for merged_range in merged_ranges:
        #unmerge all cell that >= insert_row_idx
        if insert_row_idx <= merged_range.min_row:
            sheet.unmerge_cells(str(merged_range))
        # Shift cells down
    sheet.insert_rows(insert_row_idx)

    # Insert data
    sheet.cell(row=insert_row_idx, column=1, value=data['Supplier'])
    sheet.cell(row=insert_row_idx, column=2, value=data['Remain'])
    sheet.cell(row=insert_row_idx, column=3, value=data['Description'])
    sheet.cell(row=insert_row_idx, column=4, value=data['Total'])
    sheet.cell(row=insert_row_idx, column=5, value=data['Payment'])
    sheet.cell(row=insert_row_idx, column=6, value=data['Remain'])
    sheet.cell(row=insert_row_idx, column=7, value=data['Project'])
    sheet.cell(row=insert_row_idx, column=8, value=data['Title'])
    sheet.cell(row=insert_row_idx, column=9, value=data['Date'])

    # Enforce wrap text in the row
    for cell in sheet[insert_row_idx]:
        cell.alignment = Alignment(wrapText=True)

    #Apply auto fit row height
    #Autofit the row height
    sheet.row_dimensions[insert_row_idx].height = None

    #border all cell
    for cell in sheet[insert_row_idx]:
        cell.border = Border(top=openpyxl.styles.Side(style='thin'), 
                            right=openpyxl.styles.Side(style='thin'), 
                            bottom=openpyxl.styles.Side(style='thin'), 
                            left=openpyxl.styles.Side(style='thin'))

    # Re-merge cells with shifting
    for merged_range in merged_ranges:
        if merged_range.min_row >= insert_row_idx:  # Only shift if range starts below insertion
            new_range = openpyxl.worksheet.cell_range.CellRange(
                min_row=merged_range.min_row + 1,
                min_col=merged_range.min_col,
                max_row=merged_range.max_row + 1,
                max_col=merged_range.max_col
            )
            sheet.merge_cells(str(new_range))
    
        # Merge with new row (if applicable)
    merged_ranges = list(sheet.merged_cells.ranges)  # Update merged ranges
    merged_temp = []
    if same_supplier:
        #get the previous merge range with insert_row_idx
        for merged_range in merged_ranges:
            if merged_range.max_row == insert_row_idx - 1: 
                # we get the right index now, now we need to do it with column A and B
                merged_temp.append(merged_range)

        if len(merged_temp) > 0:
            #this case mean this supplier is already merged before
            for merged_range in merged_temp:
                sheet.unmerge_cells(str(merged_range))
                #now merge it with the new row, value column 1 is merged_range.min_col, value column 2 is merged_range.min_col + int(data['Remain'])
                new_range = openpyxl.worksheet.cell_range.CellRange(
                    min_row=merged_range.min_row,
                    min_col=merged_range.min_col,
                    max_row=merged_range.max_row + 1,
                    max_col=merged_range.max_col
                )
                sheet.merge_cells(str(new_range))
                #update the value of the merged cell
                if merged_range.min_col == 2:
                    sheet.cell(row=merged_range.min_row, column=2, value=sheet.cell(row=merged_range.min_row, column=2).value + data['Remain']) #Column 2

           
        else:
            #let do for case that this suppliers is 1st time, and now the new row is 2nd time this supplier appear, we need to merge it
            new_range = openpyxl.worksheet.cell_range.CellRange(
                min_row=insert_row_idx - 1,
                min_col=2,
                max_row=insert_row_idx,
                max_col=2
            )
            sheet.merge_cells(str(new_range))
            sheet.cell(row=insert_row_idx - 1, column=2, value=sheet.cell(row=insert_row_idx - 1, column=2).value + data['Remain'])

def process(df_debit,df_report, workbook, sheet, supplier_positions, total_row,path_report):
    for _, row_data in df_debit.iterrows():
        row_subset = row_data[['Project', 'Supplier', 'Title']]

        # Filter df based on partial matches (handling NaNs)
        matches = df_report[
            (df_report['PROJECT'] == row_subset['Project']) &
            (df_report['SUPPLIERS'] == row_subset['Supplier']) &
            (df_report['TITLE'] == row_subset['Title']) 
        ]

        if not matches.empty: 
            index = matches.index[0] + 3  # 1-based index,1 skip row, 1 header row
            # Update existing row
            eddit_row_data(sheet, index, row_data)

        else:
            if row_data['Supplier'] in supplier_positions.values():
                #get insert_row is the last row of same supplier
                insert_row = max([key for key, value in supplier_positions.items() if value == row_data['Supplier']]) + 1

                # insert_row = supplier_positions[row_data['Supplier']] + 1  # Insert below existing

                insert_row_data(sheet, insert_row, row_data, True)
                
                # Update subsequent positions (increase key by 1 if greater than insert_row)
                # Update subsequent positions
                supplier_positions = {key + (key >= insert_row): value for key, value in supplier_positions.items()}
                #re-load the supplier_positions
                supplier_positions = get_supplier_positions(sheet)

                sheet.merged_cells = sheet.merged_cells  # Reload merged cells 
                total_row += 1 #update total_row
                # print('Insert to row:', insert_row)
                # print(supplier_positions)
                
            else:
                insert_row = total_row
                insert_row_data(sheet, insert_row, row_data)
                
                # Update subsequent positions
                supplier_positions = {key + (key >= insert_row): value for key, value in supplier_positions.items()}
                #re-load the supplier_positions
                supplier_positions = get_supplier_positions(sheet)

                sheet.merged_cells = sheet.merged_cells  # Reload merged cells
                total_row += 1 #update total_row
                # print('Insert to row:', insert_row)
                # print(supplier_positions)
    workbook.save(filename=path_report)
    # print('Done')

def debit_import(path_data, path_report):
    sheet_debits = [sheet.title for sheet in openpyxl.load_workbook(path_data).worksheets]
    #remove sheet name 'template'
    sheet_debits.remove('template')
    for sheet_debit in sheet_debits:
        df_report = pd.read_excel(path_report, sheet_name="DEBIT", engine='openpyxl',skiprows=1, usecols="A:J")
        df_debit = extract_data(path_data, sheet_debit)
        workbook, sheet, supplier_positions, total_row = append_data(path_report)
        process(df_debit,df_report, workbook, sheet, supplier_positions, total_row,path_report)


# def main():
#     path = "./January.xlsx"
#     sheet_debits = [sheet.title for sheet in openpyxl.load_workbook(path).worksheets]
#     #remove sheet name 'template'
#     sheet_debits.remove('template')
#     path_report = "./report_v0.0.2.xlsx"
#     for sheet_debit in sheet_debits:
#         df_report = pd.read_excel(path_report, sheet_name="DEBIT", engine='openpyxl',skiprows=1, usecols="A:J")
#         df_debit = extract_data(path, sheet_debit)
#         workbook, sheet, supplier_positions, total_row = append_data(path_report)
#         process(df_debit,df_report, workbook, sheet, supplier_positions, total_row,path_report)

def debit(xlsx_report:str, xlsx_monthly:str, sheets: list):
    for sheet_debit in sheets:
        try:
            df_report = pd.read_excel(xlsx_report, sheet_name="DEBIT", engine='openpyxl',skiprows=1, usecols="A:J")
            df_debit = extract_data(xlsx_monthly, sheet_debit)
            workbook, sheet, supplier_positions, total_row = append_data(xlsx_report)
            process(df_debit, df_report, workbook, sheet, supplier_positions, total_row, xlsx_report)
        except Exception as e:
            logfile(f'DEBIT [{datetime.datetime.now()}] Error:{e}')
    # %%
if __name__ == "__main__":
    # main()
    debit(xlsx_report, xlsx_data, date_sheets)



