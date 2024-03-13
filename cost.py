import pandas as pd
from excelautopd import get_data
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
import warnings
import re
warnings.filterwarnings("ignore")
# from configs import xlsx_report, xlsx_data, date_sheets

def get_merged_cells(worksheet: Worksheet):
    merged_cells = []
    for merge_cell in worksheet.merged_cells:
        merged_cells.append((
        merge_cell.min_row, 
        merge_cell.min_col, 
        merge_cell.max_row, 
        merge_cell.max_col
        ))
    return merged_cells

def unmerge_cells(worksheet: Worksheet, merged_cells: list):
    for merge_cell in merged_cells:
        worksheet.unmerge_cells(
            start_row=merge_cell[0],
            start_column=merge_cell[1],
            end_row=merge_cell[2],
            end_column=merge_cell[3]
        )


def extract_data(path, sheet_name):
    df = get_data(path, sheet_name)
    df_debit = df[['Project','Supplier', 'Description', 'Total', 'Payment', 'Remain']].copy()
    
    df_debit.loc[:, 'Title'] = df['Description'].apply(lambda x: x.split('\n')[0])
    df_debit.loc[:,'Date'] = sheet_name

    return df_debit

def get_list_day_payment(test_data,sheet_name,num):
    list_day_payment = []
    list_description = test_data.split("\n")
    if 'total' in list_description[-1]:
        list_day_payment.append(sheet_name)
    elif sum([True for i in num if i not in test_data])==len(num) :
        list_day_payment.append(sheet_name)
    else:
        for text in list_description:
            if any(i in text for i in num):
                re_t = re.findall(r'\(\d{2}/\d{2}/\d{4}\)', text)
                if len(re_t) > 0:
                    day_payment = re_t[0].replace("(","").replace(")","")
                else:
                    day_payment=sheet_name
                list_day_payment.append(day_payment)
    # remove duplicate
    # list_day_payment = list(set(list_day_payment))
    return ", ".join(list_day_payment)

def process_step(path_cost, path_daily, sheet_name_daily ,sheet_name_cost = 'COST'):
    df_bkcp = pd.read_excel(path_cost,sheet_name=sheet_name_cost,skiprows=4)
    df_bkcp_project = df_bkcp['Projects'].unique().tolist()[1:]
    df_bkcp_Supplier = df_bkcp['Sub-contractor/ Supplier'].unique().tolist()[1:]
    df_bkcp_Content = df_bkcp['Content.1'].unique().tolist()[1:]
    
    df = extract_data(path_daily,sheet_name_daily)
    projects_name_bkcp = df_bkcp['Projects'].unique().tolist()
    projects_name_import = df['Project'].unique().tolist()

    num = ['1st','2nd','3rd','31st'] + [str(i)+'th' for i in range(4,31)]
    
    for i,row in df.iterrows():
        data_new= {i:'' for i in df_bkcp.columns}
        project= row['Project'].split("-",1)[-1].strip()
        suplier = row['Supplier']
        description = row['Description']
        total = row['Total']
        remain = row['Remain']
        subtotal = row['Payment']
        data_new['Projects']=project
        data_new['Original Contract Value'] = 0
        data_new['Liquid Contract Value'] = 0
        data_new['Sub-contractor/ Supplier']=suplier
        data_new['Content.1']=description
        data_new['Contract Value']=int(total)
        data_new['Payment']=int(subtotal)
        data_new['Remain payment']=int(remain)
        data_new['Date payment']=get_list_day_payment(description,sheet_name_daily,num)
        
        # if project in df_bkcp_project or suplier in df_bkcp_Supplier or description in df_bkcp_Content:
        #     continue
        # print(data_new)
        if project not in projects_name_bkcp[1:]:
            df_new = pd.DataFrame(data_new,index=[0])
            df_bkcp = pd.concat([df_bkcp,df_new],ignore_index=True)
        else:
            index_insert = 0
            check = False
            for i,row_old_data in enumerate(df_bkcp.iterrows()):
                if i ==0:
                    continue
                old_data = row_old_data[1]['Content.1']
                old_project = row_old_data[1]['Projects']
                old_suplier = row_old_data[1]['Sub-contractor/ Supplier']
                old_date_of_payment = row_old_data[1]['Date payment']
                
                if 'remain' in old_data or 'Remain' in old_data:
                    datax= old_data.split('remain')
                    if len(datax) == 1:
                        datax = old_data.split('Remain')
                    description_old = datax[0]
                    check = True
                if check and description_old.strip() in data_new['Content.1'] and data_new['Projects'] == row_old_data[1]['Projects'] and data_new['Sub-contractor/ Supplier'] == row_old_data[1]['Sub-contractor/ Supplier']:
                    index_insert = i
                    break
            if index_insert == 0:
                df1 = df_bkcp[df_bkcp['Projects']==project]
                index_end = df1.index[-1] + 1
                df_bkcp = pd.concat([df_bkcp.iloc[:index_end],pd.DataFrame(data_new,index=[0]),df_bkcp.iloc[index_end:]]).reset_index(drop=True)
            else:
                df_bkcp.loc[index_insert,:]= data_new
    projects_name_bkcp = df_bkcp['Projects'].unique().tolist()
    rows_merge = {i:{"start_row":0,"end_row":0} for i in projects_name_bkcp[1:]}
    for project in projects_name_bkcp[1:]:
        sum_contract_value = df_bkcp[df_bkcp['Projects']==project]['Contract Value'].sum()
        df_bkcp.loc[df_bkcp['Projects']==project,'Total sub- contract'] = sum_contract_value
        sum_original_contract_value = df_bkcp[df_bkcp['Projects']==project]['Original Contract Value'].sum()
        df_bkcp.loc[df_bkcp['Projects']==project,'Original Contract Value'] = sum_original_contract_value
        sum_liquid_contract_value = df_bkcp[df_bkcp['Projects']==project]['Liquid Contract Value'].sum()
        df_bkcp.loc[df_bkcp['Projects']==project,'Liquid Contract Value'] = sum_liquid_contract_value
        
        start_row = df_bkcp[df_bkcp['Projects']==project].index[0]
        end_row = df_bkcp[df_bkcp['Projects']==project].index[-1]
        rows_merge[project]['start_row'] = start_row
        rows_merge[project]['end_row'] = end_row
    
    # total_sub_contract = df_bkcp['Total sub- contract'].sum()
    
    workbook = load_workbook(path_cost)
    sheet = workbook[sheet_name_cost]
    merged_cells = get_merged_cells(sheet)
    merged_cells.remove((1, 1, 2, 17))
    unmerge_cells(sheet,merged_cells)
    max_row = 6
    for i in range(len(sheet['A'])-max_row+1):
        sheet.delete_rows(max_row+1)
        
    for i,row in df_bkcp.iterrows():
        if i == 0:
            continue
        data = row.values
        sheet.insert_rows(max_row+i)
        sheet.cell(row=max_row+i,column=2).value = data[1]
        sheet.cell(row=max_row+i,column=3).value = data[2]
        sheet.cell(row=max_row+i,column=7).value = data[6]
        sheet.cell(row=max_row+i,column=8).value = data[7]
        sheet.cell(row=max_row+i,column=9).value = data[8]
        sheet.cell(row=max_row+i,column=11).value = data[10]
        sheet.cell(row=max_row+i,column=12).value = data[11]
        sheet.cell(row=max_row+i,column=13).value = int(data[12])
        sheet.cell(row=max_row+i,column=14).value = int(data[13])
        sheet.cell(row=max_row+i,column=15).value = int(data[14])
        sheet.cell(row=max_row+i,column=16).value = data[15]
        
    list_color_hex = ['FFFFCC', 'FFFF99', 'FFFF66', 'FFFF33', 'FFFF00', 'CCFFFF', 'CCFFCC', 'CCFF99', 'CCFF66', 'CCFF33', 'CCFF00', '99FFFF', '99FFCC', '99FF99', '99FF66', '99FF33', '99FF00', '66FFFF', '66FFCC', '66FF99', '66FF66', '66FF33', '66FF00', '33FFFF', '33FFCC', '33FF99', '33FF66', '33FF33', '33FF00', '00FFFF', '00FFCC', '00FF99', '00FF66', '00FF33', '00FF00', 'FFCCFF', 'FFCCCC', 'FFCC99', 'FFCC66', 'FFCC33', 'FFCC00', 'CCCCFF', 'CCCCCC', 'CCCC99', 'CCCC66', 'CCCC33', 'CCCC00', '99CCFF', '99CCCC', '99CC99', '99CC66', '99CC33', '99CC00', '66CCFF', '66CCCC', '66CC99', '66CC66', '66CC33', '66CC00', '33CCFF', '33CCCC', '33CC99', '33CC66', '33CC33', '33CC00', '00CCFF', '00CCCC', '33CC66', '33CC33', '00CC99', '00CC66']
    
    list_top=[]
    list_bottom=[]
    for key, value in rows_merge.items():
        list_top.append(value['start_row']-1)
        list_bottom.append(value['end_row'] -1 )

    list_col = []
    for index_col,project in enumerate(projects_name_bkcp[1:]):
        start_row = rows_merge[project]['start_row']
        end_row = rows_merge[project]['end_row']
        list_col.extend([list_color_hex[index_col] for i in range(start_row,end_row+1)])
        
    style = Font(name="Times New Roman", size=12)
    thin_border = Border(left=Side(style='thin'), 
                            right=Side(style='thin'), 
                            top=Side(style='thin'), 
                            bottom=Side(style='thin'))

    thick_top_border = Border(top=Side(style='thick'), 
                            left=Side(style='thin'), 
                            right=Side(style='thin'), 
                            bottom=Side(style='thin'))

    thick_bottom_border = Border(top=Side(style='thin'), 
                                left=Side(style='thin'), 
                                right=Side(style='thin'), 
                            bottom=Side(style='thick'))
    start_row = 7
    end_row = len(sheet["A"])
    cell_range = f'A{start_row}:Q{end_row}'
    
    for index_color,row in enumerate(sheet[cell_range]):
        if index_color in list_top:
            choose_border = thick_top_border
        elif index_color in list_bottom:
            choose_border = thick_bottom_border
        else:
            choose_border = thin_border
        for cell in (row):
            cell.font = style
            cell.border = choose_border
            cell.fill = PatternFill(start_color=list_col[index_color], end_color=list_col[index_color], fill_type="solid")
    
    alignment_style = Alignment(horizontal='center', vertical='center',wrapText=True)
    start_row = 7
    end_row = len(sheet["A"])  # Assuming column A has the data
    list_cell_range = [f'P{start_row}:P{end_row}', f'G{start_row}:I{end_row}', f'K{start_row}:K{end_row}']
    for cell_range in list_cell_range:
        for row in sheet[cell_range]:
            for cell in row:
                cell.alignment  = alignment_style
    
    alignment_style = Alignment(vertical='center',wrapText=True)
    list_cell_range = [f'B{start_row}:B{end_row}', f'L{start_row}:L{end_row}', f'M{start_row}:O{end_row}']
    for cell_range in list_cell_range:
        for row in sheet[cell_range]:
            for cell in row:
                cell.alignment  = alignment_style
    
    for project in projects_name_bkcp[1:]:
        start_row = rows_merge[project]['start_row'] + 6
        end_row = rows_merge[project]['end_row'] + 6
        sheet.merge_cells(start_row=start_row, start_column=7, end_row=end_row, end_column=7)
        sheet.merge_cells(start_row=start_row, start_column=8, end_row=end_row, end_column=8)
        sheet.merge_cells(start_row=start_row, start_column=9, end_row=end_row, end_column=9)
        
    workbook.save(path_cost)
    print(f"Done sheet {sheet_name_daily}!")

# def process(path_cost, path_daily):
#     workbook_daily = load_workbook(filename=path_daily)
#     # get all sheet names
#     sheet_name_dailys = workbook_daily.sheetnames
#     sheet_name_dailys.remove('template')
#     for sheet_name_daily in sheet_name_dailys:
#         process_step(path_cost, path_daily, sheet_name_daily)
#     print("Done!")

def cost(xlsx_report, xlsx_data, date_sheets):
    for sheet_name_daily in date_sheets:
        print(sheet_name_daily)
        process_step(xlsx_report, xlsx_data, sheet_name_daily)

if __name__ == "__main__":
    cost(xlsx_report, xlsx_data, date_sheets)